"""Tests for services/session_service.py.

Covers import_excel, export_excel, and the formula-injection sanitiser.
No Qt, no network.
"""
import json
from pathlib import Path
from typing import Dict, List
from unittest.mock import MagicMock

import pytest

pytest.importorskip("openpyxl")

from services.session_service import import_excel, export_excel, _safe_cell


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def make_data_mgr(meta: Dict = None):
    dm = MagicMock()
    dm.get_meta.return_value = meta or {}
    return dm


CATEGORIES = [{"name": "Säck", "description": "En säck"}, {"name": "Hink", "description": ""}]
CAT_KNOWLEDGE = {"Säck": "VISUELLA KRAV:\n- Säckform"}
CAT_EXAMPLE = {"Säck": ["10000", "10001"]}

RESULTS = [
    {"article_number": "10000", "category": "Säck", "url": "http://ex.com/a.jpg", "bolag": "AB"},
    {"article_number": "10001", "category": "Hink", "url": "http://ex.com/b.jpg", "bolag": "CD"},
]
CSV_DATA = [
    {"article_number": "10000", "url": "http://ex.com/a.jpg", "bolag": "AB", "img_path": None},
    {"article_number": "10001", "url": "http://ex.com/b.jpg", "bolag": "CD", "img_path": None},
    {"article_number": "99999", "url": "http://ex.com/c.jpg", "bolag": "",  "img_path": None},
]


# ---------------------------------------------------------------------------
# _safe_cell
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("value,expected", [
    ("=SUM(A1)", "'=SUM(A1)"),
    ("+1", "'+1"),
    ("-1", "'-1"),
    ("@macro", "'@macro"),
    ("normal text", "normal text"),
    ("", ""),
    ("10000", "10000"),
])
def test_safe_cell_formula_injection(value, expected):
    assert _safe_cell(value) == expected


def test_safe_cell_none_passthrough():
    assert _safe_cell(None) is None


# ---------------------------------------------------------------------------
# export_excel
# ---------------------------------------------------------------------------

def test_export_creates_file(tmp_path):
    out = str(tmp_path / "out.xlsx")
    export_excel(out, "TestSession", CATEGORIES, RESULTS, CSV_DATA,
                 CAT_KNOWLEDGE, CAT_EXAMPLE, data_mgr=None)
    assert Path(out).exists()
    assert Path(out).stat().st_size > 0


def test_export_creates_all_sheets(tmp_path):
    import openpyxl
    out = str(tmp_path / "out.xlsx")
    export_excel(out, "TestSession", CATEGORIES, RESULTS, CSV_DATA,
                 CAT_KNOWLEDGE, CAT_EXAMPLE, data_mgr=None)
    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    assert "Resultat" in wb.sheetnames
    assert "Session" in wb.sheetnames
    assert "Kategorianalys" in wb.sheetnames
    assert "Oklassificerade" in wb.sheetnames
    wb.close()


def test_export_no_oklassificerade_when_all_classified(tmp_path):
    """No Oklassificerade sheet if csv_data == results."""
    import openpyxl
    out = str(tmp_path / "out.xlsx")
    csv_only_classified = [
        {"article_number": r["article_number"], "url": r["url"],
         "bolag": r["bolag"], "img_path": None}
        for r in RESULTS
    ]
    export_excel(out, "S", CATEGORIES, RESULTS, csv_only_classified,
                 {}, {}, data_mgr=None)
    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    assert "Oklassificerade" not in wb.sheetnames
    wb.close()


def test_export_no_kategorianalys_when_empty_knowledge(tmp_path):
    import openpyxl
    out = str(tmp_path / "out.xlsx")
    export_excel(out, "S", CATEGORIES, RESULTS, CSV_DATA, {}, {}, data_mgr=None)
    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    assert "Kategorianalys" not in wb.sheetnames
    wb.close()


def test_export_session_sheet_contains_json(tmp_path):
    import openpyxl
    out = str(tmp_path / "out.xlsx")
    export_excel(out, "MyTest", CATEGORIES, RESULTS, CSV_DATA,
                 CAT_KNOWLEDGE, CAT_EXAMPLE, data_mgr=None)
    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    ws = wb["Session"]
    kv = {str(row[0]): str(row[1]) for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
    assert kv["test_name"] == "MyTest"
    cats_loaded = json.loads(kv["categories_json"])
    assert any(c["name"] == "Säck" for c in cats_loaded)
    wb.close()


def test_export_resultat_contains_article_numbers(tmp_path):
    import openpyxl
    out = str(tmp_path / "out.xlsx")
    export_excel(out, "S", CATEGORIES, RESULTS, CSV_DATA,
                 {}, {}, data_mgr=None)
    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    ws = wb["Resultat"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    art_nums = [str(r[0]) for r in rows if r[0]]
    assert "10000" in art_nums
    assert "10001" in art_nums
    wb.close()


def test_export_oklassificerade_contains_unclassified(tmp_path):
    import openpyxl
    out = str(tmp_path / "out.xlsx")
    export_excel(out, "S", CATEGORIES, RESULTS, CSV_DATA,
                 {}, {}, data_mgr=None)
    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    ws = wb["Oklassificerade"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    art_nums = [str(r[0]) for r in rows if r[0]]
    assert "99999" in art_nums
    # classified articles should not be in Oklassificerade
    assert "10000" not in art_nums
    wb.close()


def test_export_uses_data_mgr_metadata(tmp_path):
    import openpyxl
    out = str(tmp_path / "out.xlsx")
    dm = make_data_mgr({"beskrivning": "Foder 20kg", "langd": "300"})
    export_excel(out, "S", CATEGORIES, RESULTS[:1], CSV_DATA[:1],
                 {}, {}, data_mgr=dm)
    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    ws = wb["Resultat"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    h = {v: i for i, v in enumerate(headers) if v}
    assert rows[0][h["Beskrivning"]] == "Foder 20kg"
    assert rows[0][h["Längd (mm)"]] == "300"
    wb.close()


def test_export_sanitises_formula_in_article_number(tmp_path):
    import openpyxl
    results_with_formula = [
        {"article_number": "=CMD|'/c calc'!A0", "category": "Säck",
         "url": "", "bolag": ""}
    ]
    csv_formula = [{"article_number": "=CMD|'/c calc'!A0", "url": "", "bolag": "", "img_path": None}]
    out = str(tmp_path / "out.xlsx")
    export_excel(out, "S", CATEGORIES, results_with_formula, csv_formula,
                 {}, {}, data_mgr=None)
    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    ws = wb["Resultat"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    cell_value = str(rows[0][0])
    assert not cell_value.startswith("=")
    wb.close()


# ---------------------------------------------------------------------------
# import_excel
# ---------------------------------------------------------------------------

def _write_session(tmp_path, *, test_name="TestImport",
                   categories=None, cat_knowledge=None,
                   cat_example_articles=None,
                   results=None, unclassified=None) -> str:
    """Helper: write a minimal valid session Excel and return its path."""
    import openpyxl
    categories = categories or CATEGORIES
    cat_knowledge = cat_knowledge or CAT_KNOWLEDGE
    cat_example_articles = cat_example_articles or CAT_EXAMPLE
    results = results or RESULTS
    unclassified = unclassified or []

    path = str(tmp_path / "session.xlsx")
    export_excel(path, test_name, categories, results,
                 [{"article_number": r["article_number"], "url": r["url"],
                   "bolag": r["bolag"], "img_path": None} for r in results]
                 + [{"article_number": u["article_number"], "url": u.get("url", ""),
                     "bolag": u.get("bolag", ""), "img_path": None} for u in unclassified],
                 cat_knowledge, cat_example_articles, data_mgr=None)
    return path


def test_import_roundtrip_test_name(tmp_path):
    path = _write_session(tmp_path, test_name="RoundTrip")
    session = import_excel(path)
    assert session["test_name"] == "RoundTrip"


def test_import_roundtrip_categories(tmp_path):
    path = _write_session(tmp_path)
    session = import_excel(path)
    names = [c["name"] for c in session["categories"]]
    assert "Säck" in names
    assert "Hink" in names


def test_import_roundtrip_cat_knowledge(tmp_path):
    path = _write_session(tmp_path)
    session = import_excel(path)
    assert "Säck" in session["cat_knowledge"]
    assert "VISUELLA KRAV" in session["cat_knowledge"]["Säck"]


def test_import_roundtrip_results(tmp_path):
    path = _write_session(tmp_path)
    session = import_excel(path)
    art_nums = [r["article_number"] for r in session["results"]]
    assert "10000" in art_nums
    assert "10001" in art_nums


def test_import_roundtrip_category_value(tmp_path):
    path = _write_session(tmp_path)
    session = import_excel(path)
    by_art = {r["article_number"]: r["category"] for r in session["results"]}
    assert by_art["10000"] == "Säck"
    assert by_art["10001"] == "Hink"


def test_import_roundtrip_cat_example_articles(tmp_path):
    path = _write_session(tmp_path)
    session = import_excel(path)
    assert "Säck" in session["cat_example_articles"]
    assert "10000" in session["cat_example_articles"]["Säck"]


def test_import_csv_data_includes_unclassified(tmp_path):
    path = _write_session(tmp_path,
                          unclassified=[{"article_number": "99999", "url": "", "bolag": ""}])
    session = import_excel(path)
    csv_arts = [r["article_number"] for r in session["csv_data"]]
    assert "99999" in csv_arts


def test_import_images_list_length_matches_results(tmp_path):
    path = _write_session(tmp_path)
    session = import_excel(path)
    assert len(session["images"]) == len(session["results"])
    assert all(v is None for v in session["images"])


def test_import_missing_session_sheet_uses_filename(tmp_path):
    """File without Session sheet: test_name falls back to filename stem."""
    import openpyxl
    path = str(tmp_path / "my_session_name.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultat"
    ws.append(["Artikelnummer", "Resultat kategori", "Bild (URL)", "Bolag"])
    ws.append(["10000", "Säck", "", ""])
    wb.save(path)
    session = import_excel(path)
    assert session["test_name"] == "my_session_name"


def test_import_missing_resultat_sheet_returns_empty_results(tmp_path):
    import openpyxl
    path = str(tmp_path / "empty.xlsx")
    wb = openpyxl.Workbook()
    ws_s = wb.active
    ws_s.title = "Session"
    ws_s.append(["Nyckel", "Värde"])
    ws_s.append(["test_name", "EmptyTest"])
    ws_s.append(["categories_json", "[]"])
    wb.save(path)
    session = import_excel(path)
    assert session["results"] == []
    assert session["csv_data"] == []


def test_import_malformed_json_in_categories(tmp_path):
    import openpyxl
    path = str(tmp_path / "bad.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Session"
    ws.append(["Nyckel", "Värde"])
    ws.append(["test_name", "Bad"])
    ws.append(["categories_json", "this is not json"])
    ws.append(["cat_knowledge_json", "{broken"])
    ws.append(["cat_example_articles_json", "[[["])
    wb.save(path)
    session = import_excel(path)
    # Should not raise; should default to empty structures
    assert session["categories"] == []
    assert session["cat_knowledge"] == {}
    assert session["cat_example_articles"] == {}


def test_import_skips_empty_rows_in_resultat(tmp_path):
    import openpyxl
    path = str(tmp_path / "sparse.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultat"
    ws.append(["Artikelnummer", "Resultat kategori", "Bild (URL)", "Bolag"])
    ws.append(["10000", "Säck", "", ""])
    ws.append([None, None, None, None])   # empty row
    ws.append(["10001", "Hink", "", ""])
    wb.save(path)
    session = import_excel(path)
    assert len(session["results"]) == 2


def test_import_raises_without_openpyxl(monkeypatch):
    import services.session_service as ss
    monkeypatch.setattr(ss, "OPENPYXL_AVAILABLE", False)
    with pytest.raises(ImportError, match="openpyxl"):
        import_excel("/fake/path.xlsx")


def test_export_raises_without_openpyxl(monkeypatch, tmp_path):
    import services.session_service as ss
    monkeypatch.setattr(ss, "OPENPYXL_AVAILABLE", False)
    with pytest.raises(ImportError, match="openpyxl"):
        export_excel(str(tmp_path / "x.xlsx"), "S", [], [], [], {}, {})
