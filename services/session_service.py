"""Session import/export — pure Python, no Qt.

Reads and writes the Excel format used by the application:
  - "Resultat" sheet  : classified articles + metadata columns
  - "Oklassificerade" : unclassified articles
  - "Session"         : JSON-serialised state (categories, knowledge, …)
  - "Kategorianalys"  : human-readable AI analysis per category
"""
import json
import logging
from pathlib import Path
from typing import Dict, List, Optional

try:
    import openpyxl
    import openpyxl.utils
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from core.constants import DEFAULT_SYFTE

_logger = logging.getLogger(__name__)

# Characters that turn an Excel cell into a formula — sanitise on export.
_FORMULA_PREFIXES = ("=", "+", "-", "@")


def _safe_cell(value: str) -> str:
    """Prevent formula injection by prepending a single quote when needed."""
    if value and isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return "'" + value
    return value


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def import_excel(path: str) -> Dict:
    """Read an Excel session file and return a session dict.

    Args:
        path: Absolute path to the .xlsx file.

    Returns:
        dict with keys:
            test_name           (str)
            syfte               (str)
            categories          (list)
            cat_knowledge       (dict)
            cat_example_articles (dict)
            results             (list of {article_number, category, url, bolag})
            csv_data            (list of {article_number, url, bolag, img_path})
            categorized         (list of {article_number, category, image_path})
            images              (list of None per result row)

    Raises:
        ImportError: if openpyxl is not installed.
        OSError / openpyxl exceptions: on file read errors.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel import. Install with: pip install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # ── Session sheet ─────────────────────────────────────────────────────
    session_kv: Dict[str, str] = {}
    if "Session" in wb.sheetnames:
        for row in wb["Session"].iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] is not None:
                session_kv[str(row[0])] = str(row[1])

    test_name = session_kv.get("test_name", Path(path).stem)
    syfte = DEFAULT_SYFTE

    try:
        categories = json.loads(session_kv.get("categories_json", "[]"))
    except json.JSONDecodeError as _e:
        _logger.warning("Kunde inte tolka categories_json: %s", _e)
        categories = []

    try:
        cat_knowledge = json.loads(session_kv.get("cat_knowledge_json", "{}"))
    except json.JSONDecodeError as _e:
        _logger.warning("Kunde inte tolka cat_knowledge_json: %s", _e)
        cat_knowledge = {}

    try:
        cat_example_articles = json.loads(session_kv.get("cat_example_articles_json", "{}"))
    except json.JSONDecodeError as _e:
        _logger.warning("Kunde inte tolka cat_example_articles_json: %s", _e)
        cat_example_articles = {}

    # ── Resultat sheet ────────────────────────────────────────────────────
    results: List[Dict] = []
    csv_data: List[Dict] = []
    categorized: List[Dict] = []
    images: List = []

    if "Resultat" in wb.sheetnames:
        ws = wb["Resultat"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        h = {str(v).strip(): i for i, v in enumerate(headers) if v}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            def _cell(key: str, default: str = "") -> str:
                idx = h.get(key)
                return str(row[idx]).strip() if idx is not None and row[idx] is not None else default

            art = _cell("Artikelnummer")
            cat = _cell("Resultat kategori")
            url = _cell("Bild (URL)")
            bolag = _cell("Bolag", "")
            if not art:
                continue
            results.append({"article_number": art, "category": cat,
                             "url": url, "bolag": bolag})
            csv_data.append({"article_number": art, "url": url,
                              "bolag": bolag, "img_path": None})
            categorized.append({"article_number": art, "category": cat,
                                 "image_path": ""})
            images.append(None)

    # ── Oklassificerade sheet ─────────────────────────────────────────────
    if "Oklassificerade" in wb.sheetnames:
        ws_u = wb["Oklassificerade"]
        u_headers = [c.value for c in next(ws_u.iter_rows(min_row=1, max_row=1))]
        u_h = {str(v).strip(): i for i, v in enumerate(u_headers) if v}

        for row in ws_u.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            def _ucell(key: str, default: str = "") -> str:
                idx = u_h.get(key)
                return str(row[idx]).strip() if idx is not None and row[idx] is not None else default

            art = _ucell("Artikelnummer")
            url = _ucell("Bild (URL)")
            bolag = _ucell("Bolag", "")
            if not art:
                continue
            csv_data.append({"article_number": art, "url": url,
                              "bolag": bolag, "img_path": None})

    wb.close()

    return {
        "test_name": test_name,
        "syfte": syfte,
        "categories": categories,
        "cat_knowledge": cat_knowledge,
        "cat_example_articles": cat_example_articles,
        "results": results,
        "csv_data": csv_data,
        "categorized": categorized,
        "images": images,
    }


def export_excel(
    path: str,
    test_name: str,
    categories: List[Dict],
    results: List[Dict],
    csv_data: List[Dict],
    cat_knowledge: Dict[str, str],
    cat_example_articles: Dict[str, List[str]],
    data_mgr=None,
    syfte: str = DEFAULT_SYFTE,
) -> None:
    """Write classification results to an Excel workbook.

    Args:
        path:                 Output path (will be overwritten).
        test_name:            Session / test name.
        categories:           List of {name, description} dicts.
        results:              List of {article_number, category, url, bolag, …}.
        csv_data:             All articles including unclassified ones.
        cat_knowledge:        {cat_name: knowledge_text} AI analysis.
        cat_example_articles: {cat_name: [article_number, …]}.
        data_mgr:             Optional DataManager for metadata lookup.
        syfte:                Classification purpose string.

    Raises:
        ImportError: if openpyxl is not installed.
        OSError: on write errors.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultat"

    headers = [
        "Artikelnummer", "Resultat kategori", "Huvudkategori",
        "Beskrivning", "Längd (mm)", "Bredd (mm)", "Höjd (mm)",
        "Volym", "Vikt brutto (kg)", "Vikt netto (kg)",
        "Robot (Y/N)", "StoreQuantity", "Bild (URL)",
    ]
    ws.append(headers)

    for row in results:
        art = str(row.get("article_number", ""))
        if data_mgr is not None:
            meta = data_mgr.get_meta(art, row.get("bolag", "")) or {}
        else:
            meta = {}
        ws.append([
            _safe_cell(art),
            _safe_cell(row.get("category", "")),
            meta.get("huvudkategori", ""),
            meta.get("beskrivning", ""),
            meta.get("langd", ""),
            meta.get("bredd", ""),
            meta.get("hojd", ""),
            meta.get("volym", ""),
            meta.get("vikt_brutto", ""),
            meta.get("vikt_netto", ""),
            meta.get("robot", ""),
            meta.get("store_quantity", ""),
            row.get("url", ""),
        ])

    col_widths = [20, 25, 25, 40, 12, 12, 12, 12, 18, 18, 12, 15, 60]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Oklassificerade sheet ─────────────────────────────────────────────
    classified_nums = {str(r.get("article_number", "")) for r in results}
    unclassified = [
        row for row in csv_data
        if str(row.get("article_number", "")) not in classified_nums
    ]
    if unclassified:
        ws_u = wb.create_sheet("Oklassificerade")
        ws_u.append(["Artikelnummer", "Bild (URL)", "Bolag"])
        for row in unclassified:
            ws_u.append([
                _safe_cell(str(row.get("article_number", ""))),
                row.get("url", ""),
                row.get("bolag", ""),
            ])
        ws_u.column_dimensions["A"].width = 20
        ws_u.column_dimensions["B"].width = 60
        ws_u.column_dimensions["C"].width = 15

    # ── Session sheet ─────────────────────────────────────────────────────
    ws_s = wb.create_sheet("Session")
    ws_s.append(["Nyckel", "Värde"])
    ws_s.append(["test_name", test_name])
    ws_s.append(["syfte", syfte])
    ws_s.append(["categories_json", json.dumps(categories, ensure_ascii=False)])
    if cat_knowledge:
        ws_s.append(["cat_knowledge_json", json.dumps(cat_knowledge, ensure_ascii=False)])
    if cat_example_articles:
        ws_s.append(["cat_example_articles_json",
                     json.dumps(cat_example_articles, ensure_ascii=False)])
    ws_s.column_dimensions["A"].width = 30
    ws_s.column_dimensions["B"].width = 80

    # ── Kategorianalys sheet ──────────────────────────────────────────────
    if cat_knowledge:
        ws_k = wb.create_sheet("Kategorianalys")
        ws_k.append(["Kategori", "AI-analys", "Exempelartiklar"])
        for cat_name, knowledge in cat_knowledge.items():
            example_arts = cat_example_articles.get(cat_name, [])
            ws_k.append([cat_name, knowledge, ", ".join(example_arts)])
        ws_k.column_dimensions["A"].width = 25
        ws_k.column_dimensions["B"].width = 80
        ws_k.column_dimensions["C"].width = 40

    wb.save(path)
