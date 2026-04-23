"""desktop/app.py — MainApp (QMainWindow) and main() entry point."""
import csv
import json
import logging
import os
import random
import shutil
import sys
import tempfile
import webbrowser
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from PyQt6.QtCore import QProcess, Qt, QTimer
from PyQt6.QtGui import QAction
from PyQt6.QtWidgets import (
    QApplication, QButtonGroup, QCheckBox, QDialog, QDialogButtonBox,
    QFileDialog, QFrame, QHBoxLayout, QLabel, QLineEdit, QMainWindow,
    QMessageBox, QProgressDialog, QPushButton, QRadioButton, QStackedWidget,
    QVBoxLayout, QWidget,
)

from core.app_info import APP_NAME, APP_VERSION, GITHUB_RELEASES_URL
from core.constants import (
    AI_JOB_MIN_PER_CAT, DEFAULT_AI_URL, DEFAULT_EXTERNAL_PROVIDERS,
    DEFAULT_MODEL, DEFAULT_SYFTE,
)
from core.data_manager import DataManager
from desktop.screens.ai_job_screen import AIJobScreen
from desktop.screens.ai_settings_screen import AISettingsScreen
from desktop.screens.classify_screen import ClassifyScreen
from desktop.screens.done_screen import DoneScreen
from desktop.screens.filter_screen import FilterScreen
from desktop.screens.setup_screen import SetupScreen
from desktop.screens.source_screen import SourceScreen
from desktop.workers.image_downloader import ImageDownloader
from desktop.workers.update_worker import UpdateCheckWorker, UpdateDownloadWorker
from desktop.widgets.helpers import mk_btn

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

_logger = logging.getLogger(__name__)

SILENT_UPDATE_ARGS = [
    "/VERYSILENT",
    "/SUPPRESSMSGBOXES",
    "/NORESTART",
    "/CLOSEAPPLICATIONS",
    "/FORCECLOSEAPPLICATIONS",
]

STYLE = """
QMainWindow, QWidget {
    background-color: #1e1e2e;
    color: #cdd6f4;
    font-family: 'Segoe UI', Arial, sans-serif;
    font-size: 13px;
}
QLabel { color: #cdd6f4; }
QLineEdit, QTextEdit {
    background-color: #313244;
    border: 1px solid #45475a;
    border-radius: 6px;
    color: #cdd6f4;
    padding: 5px 10px;
}
QLineEdit:focus, QTextEdit:focus { border: 1px solid #89b4fa; }
QPushButton {
    border-radius: 6px;
    padding: 8px 16px;
    font-weight: bold;
    border: none;
}
QScrollArea { border: none; }
QCheckBox { color: #cdd6f4; }
QMessageBox { background-color: #1e1e2e; }
"""


class MainApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_NAME)
        self.resize(1000, 700)
        self.setMinimumSize(820, 600)
        self.setStyleSheet(STYLE)
        self._setup_menu()

        # ── Session state
        self.test_name     = ""
        self.syfte         = ""
        self.categories: List[Dict] = []
        self.images: List[Optional[Path]] = []
        self.current_index = 0
        self.csv_data:      List[Dict] = []
        self.results:       List[Dict] = []
        self.temp_dir:      Optional[str] = None
        self.categorized:   List[Dict] = []
        self._pending_rows: List[Dict] = []

        # ── AI state
        self.ai_settings: Dict = {}
        self.ai_enabled   = False
        self.cat_knowledge: Dict[str, str] = {}
        self.cat_example_articles: Dict[str, List[str]] = {}

        # ── Data
        self.data_mgr = DataManager()

        # ── Download worker
        self.dl_worker:     Optional[ImageDownloader] = None
        self._ready_images: set = set()
        self._update_check_worker: Optional[UpdateCheckWorker] = None
        self._update_download_worker: Optional[UpdateDownloadWorker] = None
        self._update_progress: Optional[QProgressDialog] = None

        # ── Lazy screen references
        self._setup_scr: Optional[SetupScreen] = None
        self._src_scr: Optional[SourceScreen] = None
        self._flt_scr: Optional[FilterScreen] = None

        # ── Stack
        self.stack = QStackedWidget()
        self.setCentralWidget(self.stack)

        self._cl_scr   = ClassifyScreen()
        self._done_scr = DoneScreen()

        self.stack.addWidget(self._cl_scr)
        self.stack.addWidget(self._done_scr)

        # Landing screen is SourceScreen — user picks data source first.
        self._src_scr = self._make_source_screen()
        self.stack.addWidget(self._src_scr)

        # ── Connections
        self._cl_scr.classified.connect(self._on_classified)
        self._cl_scr.skipped.connect(self._on_skip)
        self._cl_scr.go_back.connect(self._on_go_back)
        self._cl_scr.add_category.connect(self._add_cat_during_test)
        self._cl_scr.category_renamed.connect(self._rename_cat_during_test)
        self._cl_scr.end_test.connect(self._show_done)
        self._cl_scr.run_ai_job.connect(self._on_run_ai_job_from_classify)

        self._done_scr.new_test.connect(self._on_new_test)
        self._done_scr.retest_ovrigt.connect(self._retest_ovrigt)
        self._done_scr.export_excel.connect(self._export_excel)
        self._done_scr.resume_job.connect(self._open_resumed_session)
        self._done_scr.quit_app.connect(self.close)

        self.stack.setCurrentWidget(self._src_scr)
        self.showMaximized()
        self._schedule_update_check()

    # ── helpers ────────────────────────────────────────────────────────────────

    def _setup_menu(self):
        help_menu = self.menuBar().addMenu("&Hjälp")
        update_action = QAction("Sök efter uppdateringar", self)
        update_action.triggered.connect(lambda: self._check_for_updates(manual=True))
        help_menu.addAction(update_action)

        release_action = QAction("Öppna releasesida", self)
        release_action.triggered.connect(lambda: webbrowser.open(GITHUB_RELEASES_URL))
        help_menu.addAction(release_action)

        help_menu.addSeparator()
        about_action = QAction(f"Om {APP_NAME}", self)
        about_action.triggered.connect(self._show_about_dialog)
        help_menu.addAction(about_action)

    def _show_about_dialog(self):
        QMessageBox.about(
            self,
            f"Om {APP_NAME}",
            f"{APP_NAME}\nVersion {APP_VERSION}",
        )

    def _schedule_update_check(self):
        if not self._automatic_update_checks_enabled():
            return
        QTimer.singleShot(2500, lambda: self._check_for_updates(manual=False))

    def _automatic_update_checks_enabled(self) -> bool:
        if os.environ.get("ARTIKELPLACERING_DISABLE_UPDATE_CHECK") == "1":
            return False
        # Tests construct MainApp often and must never make network calls.
        return "pytest" not in sys.modules

    def _check_for_updates(self, manual: bool = False):
        if self._update_check_worker and self._update_check_worker.isRunning():
            if manual:
                QMessageBox.information(
                    self, "Uppdatering", "Söker redan efter uppdateringar."
                )
            return

        worker = UpdateCheckWorker(APP_VERSION)
        worker.update_available.connect(
            lambda info: self._on_update_available(info, manual)
        )
        worker.no_update.connect(lambda: self._on_no_update(manual))
        worker.error.connect(lambda msg: self._on_update_error(msg, manual))
        worker.finished.connect(lambda: self._on_update_check_finished(worker))
        self._update_check_worker = worker
        worker.start()

    def _on_update_check_finished(self, worker: UpdateCheckWorker):
        if self._update_check_worker is worker:
            self._update_check_worker = None
        worker.deleteLater()

    def _on_no_update(self, manual: bool):
        if manual:
            QMessageBox.information(
                self,
                "Ingen uppdatering",
                f"Du kör senaste versionen av {APP_NAME}.",
            )

    def _on_update_error(self, message: str, manual: bool):
        if manual:
            QMessageBox.warning(
                self,
                "Kunde inte söka efter uppdatering",
                f"Kontrollera internetanslutningen och försök igen.\n\n{message}",
            )

    def _on_update_available(self, info, manual: bool):
        if not info.installer_url:
            reply = QMessageBox.question(
                self,
                "Uppdatering finns",
                (
                    f"Version {info.version} finns tillgänglig, men releasen "
                    "saknar en Setup.exe-fil. Vill du öppna releasesidan?"
                ),
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes,
            )
            if reply == QMessageBox.StandardButton.Yes:
                webbrowser.open(info.release_url)
            return

        reply = QMessageBox.question(
            self,
            "Uppdatering finns",
            (
                f"Version {info.version} finns tillgänglig.\n\n"
                "Vill du ladda ner och installera uppdateringen nu? "
                "Appen stängs automatiskt medan uppdateringen installeras."
            ),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes if manual else QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            self._download_update(info)

    def _download_update(self, info):
        if self._update_download_worker and self._update_download_worker.isRunning():
            QMessageBox.information(
                self, "Uppdatering", "Uppdateringen laddas redan ner."
            )
            return

        target_dir = Path(tempfile.gettempdir()) / APP_NAME / "updates"
        progress = QProgressDialog("Laddar ner uppdatering...", "Avbryt", 0, 100, self)
        progress.setWindowTitle("Uppdatering")
        progress.setWindowModality(Qt.WindowModality.ApplicationModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)

        worker = UpdateDownloadWorker(info, target_dir)
        progress.canceled.connect(worker.stop)
        worker.progress.connect(progress.setValue)
        worker.downloaded.connect(self._on_update_downloaded)
        worker.error.connect(self._on_update_download_error)
        worker.finished.connect(lambda: self._on_update_download_finished(worker))
        self._update_progress = progress
        self._update_download_worker = worker
        worker.start()

    def _on_update_downloaded(self, installer_path: str):
        if self._update_progress:
            self._update_progress.setValue(100)
            self._update_progress.close()
        started = QProcess.startDetached(installer_path, SILENT_UPDATE_ARGS)
        if started:
            QApplication.quit()
        else:
            QMessageBox.critical(
                self,
                "Kunde inte starta uppdatering",
                f"Installeraren kunde inte startas:\n{installer_path}",
            )

    def _on_update_download_error(self, message: str):
        if self._update_progress:
            self._update_progress.close()
        QMessageBox.warning(
            self,
            "Kunde inte ladda ner uppdatering",
            f"Försök igen senare.\n\n{message}",
        )

    def _on_update_download_finished(self, worker: UpdateDownloadWorker):
        if self._update_download_worker is worker:
            self._update_download_worker = None
        self._update_progress = None
        worker.deleteLater()

    def _push_screen(self, widget: QWidget):
        self.stack.addWidget(widget)
        self.stack.setCurrentWidget(widget)

    def _replace_top(self, new_widget: QWidget, old_widget: Optional[QWidget]):
        if old_widget and self.stack.indexOf(old_widget) >= 0:
            self.stack.removeWidget(old_widget)
            old_widget.setParent(None)
        self._push_screen(new_widget)

    # ── navigation ─────────────────────────────────────────────────────────────

    def _make_source_screen(self) -> SourceScreen:
        src = SourceScreen(len(self.data_mgr.builtin_attributes))
        src.use_builtin.connect(self._show_filter_screen)
        src.use_csv.connect(self._stage_csv)
        src.load_excel.connect(self._import_excel)
        return src

    def _show_filter_screen(self):
        flt = FilterScreen("", list(self.data_mgr.builtin_attributes), self.data_mgr)
        flt.go_next.connect(self._stage_download)
        flt.go_back.connect(lambda: self.stack.setCurrentWidget(self._src_scr))
        self._flt_scr = flt
        self._push_screen(flt)

    def _stage_csv(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Välj CSV-fil", "", "CSV-filer (*.csv);;Alla filer (*)"
        )
        if not path:
            return
        rows = self._parse_csv(path)
        if rows:
            self._pending_rows = rows
            self._show_setup_screen(rows)

    def _stage_download(self, rows: List[Dict]):
        self._pending_rows = rows
        self._show_setup_screen(rows)

    def _show_setup_screen(self, rows: List[Dict]):
        back_target = getattr(self, "_flt_scr", None) or self._src_scr
        if self._setup_scr is not None:
            self._setup_scr.cleanup()
            self.stack.removeWidget(self._setup_scr)
            self._setup_scr.setParent(None)
        scr = SetupScreen(
            rows, self.data_mgr,
            prefill_name=self.test_name,
            prefill_cats=self.categories or None,
        )
        scr.go_next.connect(self._on_setup_done)
        scr.go_back.connect(lambda: self.stack.setCurrentWidget(back_target))
        self._setup_scr = scr
        self._push_screen(scr)

    def _on_setup_done(self, name: str, syfte: str, cats: List[Dict]):
        self.test_name = name
        self.syfte     = syfte
        self.categories = [dict(c, knowledge="") for c in cats]
        if self._setup_scr:
            self._setup_scr.cleanup()
        self._download_images(self._pending_rows)

    def _show_ai_settings(self, back_target: Optional[QWidget] = None):
        if back_target is None:
            back_target = (getattr(self, "_setup_scr", None)
                           or getattr(self, "_flt_scr", None)
                           or self._src_scr)
        ai = AISettingsScreen(self.test_name)
        ai.go_next.connect(self._on_ai_done)
        ai.go_back.connect(lambda: self.stack.setCurrentWidget(back_target))
        self._ai_scr = ai
        self._push_screen(ai)

    def _on_run_ai_job_from_classify(self):
        self._show_ai_settings(back_target=self._cl_scr)

    def _on_ai_done(self, settings: Dict):
        self.ai_settings = settings
        self.ai_enabled  = bool(settings)
        pending = getattr(self, "_pending_start", None)
        self._pending_start = None
        if pending:
            pending()
        elif self.ai_enabled:
            self._run_ai_job()
        else:
            self.stack.setCurrentWidget(self._cl_scr)

    # ── image loading ──────────────────────────────────────────────────────────

    def _parse_csv(self, path: str) -> Optional[List[Dict]]:
        try:
            with open(path, newline="", encoding="utf-8-sig") as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except csv.Error:
                    dialect = csv.excel
                all_rows = list(csv.reader(f, dialect))
            url_col = None
            for row in all_rows[:5]:
                for i, cell in enumerate(row):
                    if cell.strip().lower().startswith("http"):
                        url_col = i
                        break
                if url_col is not None:
                    break
            if url_col is None:
                QMessageBox.warning(self, "Ingen URL-kolumn",
                                    "Kunde inte hitta kolumn med URL:er.")
                return None
            rows = []
            for row in all_rows:
                if len(row) <= url_col:
                    continue
                art = row[0].strip()
                url = row[url_col].strip()
                if art and url.lower().startswith("http"):
                    rows.append({"article_number": art, "url": url})
            if not rows:
                QMessageBox.warning(self, "Inga rader", "Inga giltiga rader i filen.")
                return None
            return rows
        except Exception as e:
            QMessageBox.critical(self, "CSV-fel", f"Kunde inte läsa filen:\n{e}")
            return None

    def _download_images(self, rows: List[Dict]):
        random.shuffle(rows)
        self.csv_data = [{"article_number": r["article_number"], "url": r["url"],
                          "bolag": r.get("bolag", ""), "img_path": None} for r in rows]
        self.images        = [None] * len(rows)
        self.results       = []
        self.current_index = 0
        self._ready_images = set()

        self.temp_dir = tempfile.mkdtemp(prefix="bildklassificering_")
        if self.dl_worker:
            self.dl_worker.stop()
            self.dl_worker.wait()
        self.dl_worker = ImageDownloader(rows, self.temp_dir)
        self.dl_worker.image_ready.connect(self._on_image_ready)
        self.dl_worker.start()

        self._loading_scr = self._make_loading_screen(len(rows))
        self._push_screen(self._loading_scr)

        def poll():
            if 0 in self._ready_images:
                self.stack.removeWidget(self._loading_scr)
                self._loading_scr.setParent(None)
                self._show_classify()
            else:
                QTimer.singleShot(200, poll)
        QTimer.singleShot(200, poll)

    def _make_loading_screen(self, total: int) -> QWidget:
        w = QWidget()
        w.setStyleSheet("background:#1e1e2e;")
        lay = QVBoxLayout(w)
        lay.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl = QLabel("Hämtar bilder…")
        lbl.setStyleSheet("font-size:20px; font-weight:bold;")
        lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(lbl)
        sub = QLabel(f"{total} bilder totalt — resten hämtas i bakgrunden")
        sub.setStyleSheet("color:#6c7086;")
        sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(sub)
        return w

    def _on_image_ready(self, index: int, path: str):
        self._ready_images.add(index)
        self.images[index] = Path(path)
        if index < len(self.csv_data):
            self.csv_data[index]["img_path"] = path

    def _get_meta(self, index: int) -> Optional[Dict]:
        if index >= len(self.csv_data):
            return None
        entry = self.csv_data[index]
        return self.data_mgr.get_meta(str(entry["article_number"]), entry.get("bolag", ""))

    # ── classify screen ────────────────────────────────────────────────────────

    def _show_classify(self):
        if self.current_index >= len(self.images):
            self._show_done()
            return
        if self.current_index not in self._ready_images:
            self._show_wait_screen()
            return
        img_path = self.images[self.current_index]
        if img_path is None:
            self.current_index += 1
            self._show_classify()
            return

        meta = self._get_meta(self.current_index)
        art_num_str = str(self.csv_data[self.current_index].get("article_number", ""))
        meta = dict(meta, article_number=art_num_str) if meta else {"article_number": art_num_str}
        cat_counts, threshold, ai_job_ready = self._get_threshold_data()

        prev_cat = ""
        if self.csv_data:
            art_num = str(self.csv_data[self.current_index].get("article_number", ""))
            for e in self.categorized:
                if str(e.get("article_number", "")) == art_num:
                    prev_cat = e.get("category", "")
                    break

        self._cl_scr.show_image(
            self.test_name, self.categories,
            str(img_path), meta,
            self.current_index, len(self.images),
            cat_counts, threshold, ai_job_ready,
            prev_category=prev_cat,
        )
        self.stack.setCurrentWidget(self._cl_scr)

    def _get_threshold_data(self) -> Tuple[Dict[str, int], int, bool]:
        non_ovrigt = [c["name"] for c in self.categories if c["name"] != "Övrigt"]
        if not non_ovrigt:
            return {}, 0, False
        threshold = AI_JOB_MIN_PER_CAT
        counts: Dict[str, int] = {name: 0 for name in non_ovrigt}
        for entry in self.categorized:
            cat = entry.get("category", "")
            if cat in counts:
                counts[cat] += 1
        ready = all(counts[name] >= threshold for name in non_ovrigt)
        return counts, threshold, ready

    def _show_wait_screen(self):
        w = QWidget()
        w.setStyleSheet("background:#1e1e2e;")
        lay = QVBoxLayout(w)
        lay.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl = QLabel("Väntar på nedladdning…")
        lbl.setStyleSheet("font-size:18px;")
        lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(lbl)
        sub = QLabel(f"{len(self._ready_images)} av {len(self.images)} klara")
        sub.setStyleSheet("color:#6c7086;")
        sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(sub)
        self.stack.addWidget(w)
        self.stack.setCurrentWidget(w)

        def poll():
            if self.current_index in self._ready_images:
                self.stack.removeWidget(w)
                w.setParent(None)
                self._show_classify()
            else:
                QTimer.singleShot(300, poll)
        QTimer.singleShot(300, poll)

    # ── classify logic ─────────────────────────────────────────────────────────

    def _on_classified(self, category: str):
        if self.current_index >= len(self.images):
            return
        img_path = self.images[self.current_index]
        art_num = str(self.csv_data[self.current_index]["article_number"])
        for e in self.categorized:
            if str(e.get("article_number", "")) == art_num:
                e["category"] = category
                break
        else:
            self.categorized.append({
                "image_path":     str(img_path),
                "category":       category,
                "article_number": art_num,
            })
        for r in self.results:
            if str(r.get("article_number", "")) == art_num:
                r["category"] = category
                break
        else:
            self.results.append({
                "article_number": art_num,
                "url":            self.csv_data[self.current_index]["url"],
                "category":       category,
            })
        self.current_index += 1
        self._show_classify()

    def _on_skip(self):
        self.current_index += 1
        self._show_classify()

    def _on_go_back(self):
        if self.current_index <= 0:
            return
        self.current_index -= 1
        self._show_classify()

    def _add_cat_during_test(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Ny kategori")
        dlg.setStyleSheet(STYLE)
        dlg.setMinimumWidth(400)
        lay = QVBoxLayout(dlg)
        lay.addWidget(QLabel("Kategorinamn:"))
        edit = QLineEdit()
        lay.addWidget(edit)
        lay.addWidget(QLabel("Syfte / beskrivning:"))
        desc_edit = QLineEdit()
        desc_edit.setPlaceholderText("Beskriv syftet med kategorin (valfritt)")
        lay.addWidget(desc_edit)
        if len(self.categories) >= 9:
            hint = QLabel("OBS: Fler än 9 kategorier — ingen tangent tilldelas.")
            hint.setStyleSheet("color:#fab387; font-size:11px; font-style:italic;")
            lay.addWidget(hint)
        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)
        edit.setFocus()
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        name = edit.text().strip()
        if not name:
            return
        if any(c["name"] == name for c in self.categories) or name == "Övrigt":
            QMessageBox.warning(self, "Dubblett", f'"{name}" finns redan.')
            return
        self.categories.append({"name": name, "description": desc_edit.text().strip(), "knowledge": ""})
        self._show_classify()

    def _rename_cat_during_test(self, cat_idx: int, new_name: str, new_desc: str):
        old_name = self.categories[cat_idx]["name"]
        self.categories[cat_idx]["name"] = new_name
        self.categories[cat_idx]["description"] = new_desc
        if old_name != new_name:
            for entry in self.categorized:
                if entry.get("category") == old_name:
                    entry["category"] = new_name
            for entry in self.results:
                if entry.get("category") == old_name:
                    entry["category"] = new_name
        self._show_classify()

    # ── done screen ────────────────────────────────────────────────────────────

    def _show_done(self):
        self._cleanup_workers()
        ov_count = sum(1 for r in self.results if r.get("category") == "Övrigt")
        self._done_scr.show_results(
            self.test_name, self.categories, self.current_index,
            bool(self.results), ov_count,
            results=self.results,
        )
        self.stack.setCurrentWidget(self._done_scr)

    def _on_new_test(self):
        self._cleanup_workers()
        self._cleanup_temp()
        self._reset_state()
        if self._setup_scr is not None:
            self._setup_scr.cleanup()
            self.stack.removeWidget(self._setup_scr)
            self._setup_scr.setParent(None)
            self._setup_scr = None
        if self._flt_scr is not None:
            self.stack.removeWidget(self._flt_scr)
            self._flt_scr.setParent(None)
            self._flt_scr = None
        self.stack.setCurrentWidget(self._src_scr)

    def _retest_ovrigt(self):
        ovrigt_rows = [r for r in self.results if r.get("category") == "Övrigt"]
        if not ovrigt_rows:
            QMessageBox.information(self, "Inga bilder", "Inga Övrigt-artiklar att testa om.")
            return
        art_set = {str(r["article_number"]) for r in ovrigt_rows}
        retest_data = [d for d in self.csv_data if str(d["article_number"]) in art_set]
        missing = [d for d in retest_data
                   if not d.get("img_path") or not Path(d["img_path"]).exists()]
        if missing:
            self._download_images([{"article_number": d["article_number"],
                                    "url": d["url"],
                                    "bolag": d.get("bolag", "")} for d in missing])
            return
        self.current_index = 0
        self.csv_data = retest_data
        self.images = [Path(d["img_path"]) for d in retest_data]
        self._show_classify()

    # ── AI job ─────────────────────────────────────────────────────────────────

    def _run_ai_job(self):
        if not self.ai_enabled:
            return
        if not self.categorized:
            QMessageBox.information(self, "Inga data",
                                    "Inga manuellt klassificerade artiklar att utgå från.")
            return

        scr = AIJobScreen(
            self.categories, self.categorized, self.csv_data, self.syfte,
            self.ai_settings.get("api_url", DEFAULT_AI_URL),
            self.ai_settings.get("model", DEFAULT_MODEL),
            self.ai_settings.get("compress_images", True),
            self.data_mgr, self.test_name,
            api_key=self.ai_settings.get("api_key", ""),
        )
        scr.article_added.connect(self._on_ai_article_classified)
        scr.reclassified.connect(self._on_ai_reclassified)
        scr.knowledge_updated.connect(self._on_knowledge_updated)
        scr.finished.connect(self._show_done)
        self._push_screen(scr)
        scr.start()

    def _on_ai_article_classified(self, article_number: str, category: str, url: str):
        existing = {r["article_number"] for r in self.results}
        if article_number not in existing:
            bolag = next(
                (r.get("bolag", "") for r in self.csv_data
                 if str(r.get("article_number", "")) == article_number),
                ""
            )
            self.results.append({
                "article_number": article_number,
                "category":       category,
                "url":            url,
                "bolag":          bolag,
            })

    def _on_ai_reclassified(self, article_number: str, new_category: str):
        for r in self.results:
            if r["article_number"] == article_number:
                r["category"] = new_category
                break

    def _on_knowledge_updated(self, knowledge: Dict, example_articles: Dict):
        self.cat_knowledge         = knowledge
        self.cat_example_articles  = example_articles

    def _open_resumed_session(self):
        img_by_art  = {str(r.get("article_number", "")): r.get("img_path", "")
                       for r in self.csv_data}
        art_in_cat  = {str(c.get("article_number", "")) for c in self.categorized}

        merged = list(self.categorized)
        for r in self.results:
            art = str(r.get("article_number", ""))
            if art not in art_in_cat:
                merged.append({
                    "article_number": art,
                    "category":   r.get("category", "Övrigt"),
                    "image_path": img_by_art.get(art, ""),
                    "url":        r.get("url", ""),
                    "bolag":      r.get("bolag", ""),
                })

        classified_nums = {str(c.get("article_number", "")) for c in merged}
        has_unclassified = any(
            str(row.get("article_number", "")) not in classified_nums
            for row in self.csv_data
        )

        scr = AIJobScreen(
            self.categories, merged, self.csv_data, self.syfte,
            self.ai_settings.get("api_url", DEFAULT_AI_URL),
            self.ai_settings.get("model", DEFAULT_MODEL),
            self.ai_settings.get("compress_images", True),
            self.data_mgr, self.test_name,
            api_key=self.ai_settings.get("api_key", ""),
            pre_knowledge=self.cat_knowledge if self.cat_knowledge else None,
            pre_example_articles=self.cat_example_articles if self.cat_example_articles else None,
        )
        scr.article_added.connect(self._on_ai_article_classified)
        scr.reclassified.connect(self._on_ai_reclassified)
        scr.knowledge_updated.connect(self._on_knowledge_updated)
        scr.finished.connect(self._show_done)
        self._push_screen(scr)
        if self.cat_knowledge:
            scr._cat_knowledge         = dict(self.cat_knowledge)
            scr._cat_example_articles  = dict(self.cat_example_articles)
        scr.start(skip_worker=not has_unclassified)

    # ── Excel import ───────────────────────────────────────────────────────────

    def _import_excel(self):
        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(self, "openpyxl saknas",
                                 "Installera openpyxl:\n  pip install openpyxl")
            return
        path, _ = QFileDialog.getOpenFileName(
            self, "Öppna Excel-session", "", "Excel (*.xlsx *.xls)"
        )
        if not path:
            return
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

            session: Dict = {}
            if "Session" in wb.sheetnames:
                for row in wb["Session"].iter_rows(min_row=2, values_only=True):
                    if row[0] and row[1] is not None:
                        session[str(row[0])] = str(row[1])

            test_name = session.get("test_name", Path(path).stem)
            syfte     = DEFAULT_SYFTE
            try:
                categories = json.loads(session.get("categories_json", "[]"))
            except json.JSONDecodeError as _e:
                _logger.warning("Kunde inte tolka categories_json: %s", _e)
                categories = []
            try:
                cat_knowledge = json.loads(session.get("cat_knowledge_json", "{}"))
            except json.JSONDecodeError as _e:
                _logger.warning("Kunde inte tolka cat_knowledge_json: %s", _e)
                cat_knowledge = {}
            try:
                cat_example_articles = json.loads(session.get("cat_example_articles_json", "{}"))
            except json.JSONDecodeError as _e:
                _logger.warning("Kunde inte tolka cat_example_articles_json: %s", _e)
                cat_example_articles = {}

            results: List[Dict]    = []
            csv_data: List[Dict]   = []
            categorized: List[Dict] = []
            images: List           = []

            if "Resultat" in wb.sheetnames:
                ws = wb["Resultat"]
                headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                h = {str(v).strip(): i for i, v in enumerate(headers) if v}

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue

                    def _cell(key, default=""):
                        idx = h.get(key)
                        return str(row[idx]).strip() if idx is not None and row[idx] is not None else default

                    art   = _cell("Artikelnummer")
                    cat   = _cell("Resultat kategori")
                    url   = _cell("Bild (URL)")
                    bolag = _cell("Bolag", "")
                    if not art:
                        continue
                    results.append({"article_number": art, "category": cat,
                                    "url": url, "bolag": bolag})
                    csv_data.append({"article_number": art, "url": url,
                                     "bolag": bolag, "img_path": None})
                    categorized.append({"article_number": art, "category": cat,
                                        "image_path": ""})
                    images.append(None)

            if "Oklassificerade" in wb.sheetnames:
                ws_u = wb["Oklassificerade"]
                u_headers = [c.value for c in next(ws_u.iter_rows(min_row=1, max_row=1))]
                u_h = {str(v).strip(): i for i, v in enumerate(u_headers) if v}
                for row in ws_u.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue

                    def _ucell(key, default=""):
                        idx = u_h.get(key)
                        return str(row[idx]).strip() if idx is not None and row[idx] is not None else default

                    art   = _ucell("Artikelnummer")
                    url   = _ucell("Bild (URL)")
                    bolag = _ucell("Bolag", "")
                    if not art:
                        continue
                    csv_data.append({"article_number": art, "url": url,
                                     "bolag": bolag, "img_path": None})

            wb.close()

            self._cleanup_workers()
            self._cleanup_temp()
            self._reset_state()

            self.test_name           = test_name
            self.syfte               = syfte
            self.categories          = categories
            self.csv_data            = csv_data
            self.categorized         = categorized
            self.results             = results
            self.images              = images
            self.current_index       = len(images)
            self.cat_knowledge       = cat_knowledge
            self.cat_example_articles = cat_example_articles

            if results:
                self._pending_start = self._open_resumed_session
                self._show_ai_settings()
            else:
                QMessageBox.information(self, "Tom session", "Inga resultat hittades i filen.")

        except Exception as e:
            QMessageBox.critical(self, "Fel", f"Kunde inte läsa Excel-filen:\n{e}")

    # ── Excel export ───────────────────────────────────────────────────────────

    def _export_excel(self):
        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(self, "openpyxl saknas",
                                 "Installera openpyxl:\n  pip install openpyxl")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Spara Excel", f"{self.test_name}_resultat.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultat"
        headers = [
            "Artikelnummer", "Resultat kategori", "Huvudkategori",
            "Beskrivning", "Längd (mm)", "Bredd (mm)", "Höjd (mm)",
            "Volym", "Vikt brutto (kg)", "Vikt netto (kg)",
            "Robot (Y/N)", "StoreQuantity", "Bild (URL)",
        ]
        ws.append(headers)
        for row in self.results:
            art  = str(row.get("article_number", ""))
            meta = self.data_mgr.get_meta(art, row.get("bolag", "")) or {}
            ws.append([
                art,
                row.get("category", ""),
                meta.get("huvudkategori", ""),
                meta.get("beskrivning", ""),
                meta.get("langd", ""),
                meta.get("bredd", ""),
                meta.get("hojd", ""),
                meta.get("volym", ""),
                meta.get("vikt_brutto", ""),
                meta.get("vikt_netto", ""),
                meta.get("robot", ""),
                meta.get("store_quantity", ""),
                row.get("url", ""),
            ])
        col_widths = [20, 25, 25, 40, 12, 12, 12, 12, 18, 18, 12, 15, 60]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        classified_nums = {str(r.get("article_number", "")) for r in self.results}
        unclassified = [
            row for row in self.csv_data
            if str(row.get("article_number", "")) not in classified_nums
        ]
        if unclassified:
            ws_u = wb.create_sheet("Oklassificerade")
            ws_u.append(["Artikelnummer", "Bild (URL)", "Bolag"])
            for row in unclassified:
                ws_u.append([
                    str(row.get("article_number", "")),
                    row.get("url", ""),
                    row.get("bolag", ""),
                ])
            ws_u.column_dimensions["A"].width = 20
            ws_u.column_dimensions["B"].width = 60
            ws_u.column_dimensions["C"].width = 15

        ws_s = wb.create_sheet("Session")
        ws_s.append(["Nyckel", "Värde"])
        ws_s.append(["test_name", self.test_name])
        ws_s.append(["syfte", self.syfte])
        ws_s.append(["categories_json", json.dumps(self.categories, ensure_ascii=False)])
        if self.cat_knowledge:
            ws_s.append(["cat_knowledge_json",
                         json.dumps(self.cat_knowledge, ensure_ascii=False)])
        if self.cat_example_articles:
            ws_s.append(["cat_example_articles_json",
                         json.dumps(self.cat_example_articles, ensure_ascii=False)])
        ws_s.column_dimensions["A"].width = 30
        ws_s.column_dimensions["B"].width = 80

        if self.cat_knowledge:
            ws_k = wb.create_sheet("Kategorianalys")
            ws_k.append(["Kategori", "AI-analys", "Exempelartiklar"])
            for cat_name, knowledge in self.cat_knowledge.items():
                example_arts = self.cat_example_articles.get(cat_name, [])
                ws_k.append([cat_name, knowledge, ", ".join(example_arts)])
            ws_k.column_dimensions["A"].width = 25
            ws_k.column_dimensions["B"].width = 80
            ws_k.column_dimensions["C"].width = 40

        try:
            wb.save(path)
            QMessageBox.information(self, "Exporterat", f"Sparad:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Fel", f"Kunde inte spara:\n{e}")

    # ── cleanup ────────────────────────────────────────────────────────────────

    def _cleanup_workers(self):
        if self.dl_worker:
            self.dl_worker.stop()
            self.dl_worker.wait()
            self.dl_worker = None
        if self._update_download_worker and self._update_download_worker.isRunning():
            self._update_download_worker.stop()
            self._update_download_worker.wait(3000)
        if self._update_check_worker and self._update_check_worker.isRunning():
            self._update_check_worker.wait(3000)

    def _cleanup_temp(self):
        if self.temp_dir and Path(self.temp_dir).exists():
            shutil.rmtree(self.temp_dir, ignore_errors=True)
        self.temp_dir = None

    def _reset_state(self):
        self.test_name     = ""
        self.syfte         = ""
        self.categories    = []
        self.images        = []
        self.current_index = 0
        self.csv_data      = []
        self.results       = []
        self.categorized   = []
        self.ai_settings   = {}
        self.ai_enabled    = False
        self._ready_images = set()
        self.cat_knowledge         = {}
        self.cat_example_articles  = {}

    def closeEvent(self, event):
        self._cleanup_workers()
        self._cleanup_temp()
        super().closeEvent(event)


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    win = MainApp()
    win.show()
    sys.exit(app.exec())
